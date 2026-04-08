import pandas as pd
import io
from openpyxl import load_workbook
from decimal import Decimal
from src.models import ExcelData, Totals, Calc

def get_precise_float_from_excel(workbook, sheet_name, row_idx, col_idx):
    """
    Читает число из Excel с сохранением точного количества знаков после запятой.
    Использует Decimal для точного представления и определяет количество знаков из значения.
    """
    try:
        sheet = workbook[sheet_name]
        cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)  # openpyxl использует 1-based индексацию
        
        if cell.value is None:
            return 0.0
        
        # Если значение - число
        if isinstance(cell.value, (int, float)):
            original_float = float(cell.value)
            
            # Используем Decimal для точного представления
            # Важно: преобразуем через строку для избежания потери точности
            decimal_value = Decimal(str(original_float))
            
            # Определяем количество знаков после запятой
            # Если число целое, возвращаем как float
            if decimal_value == decimal_value.to_integral_value():
                return float(decimal_value)
            
            # Для чисел с дробной частью используем Decimal для определения точности
            sign, digits, exponent = decimal_value.as_tuple()
            
            # Если exponent отрицательный, значит есть дробная часть
            if exponent < 0:
                decimal_places = abs(exponent)
                
                # Преобразуем в строку для анализа артефактов float
                str_decimal = str(decimal_value)
                
                if '.' in str_decimal:
                    fractional = str_decimal.split('.')[1]
                    
                    # Проверяем на артефакты float точности
                    # Если последние цифры - это много 9 или 0, возможно это артефакт
                    # Пробуем найти оптимальное количество знаков
                    if len(fractional) > 10:
                        # Если слишком много знаков, пробуем найти реальное количество
                        # Проверяем различные варианты округления
                        best_result = original_float
                        best_places = min(decimal_places, 10)
                        
                        for test_places in range(10, 0, -1):
                            test_rounded = round(original_float, test_places)
                            test_str = str(Decimal(str(test_rounded)))
                            
                            # Проверяем, не создали ли мы артефакт
                            if '.' in test_str:
                                test_fractional = test_str.split('.')[1]
                                # Если результат не имеет артефактов (999... или 000...), используем его
                                if not (test_fractional.endswith('999') or test_fractional.endswith('000') or len(test_fractional) > test_places + 2):
                                    best_result = test_rounded
                                    best_places = test_places
                                    break
                        
                        return best_result
                    else:
                        # Если количество знаков разумное, используем его
                        if decimal_places > 15:
                            decimal_places = 15
                        return round(original_float, decimal_places)
            
            return float(decimal_value)
        
        # Если это строка, пытаемся преобразовать
        if isinstance(cell.value, str):
            try:
                # Используем Decimal для точного преобразования
                decimal_value = Decimal(cell.value)
                # Определяем количество знаков после запятой из строки
                if '.' in cell.value:
                    fractional_part = cell.value.split('.')[1]
                    decimal_places = len(fractional_part)
                    if decimal_places > 15:
                        decimal_places = 15
                    return round(float(decimal_value), decimal_places)
                return float(decimal_value)
            except:
                return 0.0
        
        return 0.0
    except Exception as e:
        # В случае ошибки возвращаем 0.0
        return 0.0

def process_unified(file_content: bytes, CON_NUMBER: str = None) -> dict:
    
    storage = ExcelData(
        containers={},
        container_info={},
        totals=Totals(),
        calc=Calc(),
        invoice="-",
        date_invoice="-",
        sender_name="-",
        sender_address="-",
        recipient_name="-",
        recipient_address="-"
    )

    column_mapping = {
        "Unnamed: 0": "Номер",
        "Unnamed: 1": "Код ТН ВЭД",
        "Unnamed: 2": "Коммерческое описание товара",
        "Unnamed: 3": "Кол-во штук",
        "Unnamed: 4": "Кол-во мест",
        "Unnamed: 5": "Упаковка",
        "Unnamed: 6": "Нетто",
        "Unnamed: 7": "Брутто",
        "Unnamed: 8": "Валюта",
        "Unnamed: 9": "Сумма",
        "Unnamed: 10": "Номер контейнера",
        "Unnamed: 11": "Номер инвойса",
        "Unnamed: 12": "Дата инвойса",
        "Unnamed: 13": "Отправитель:",
        "Unnamed: 14": "Адрес отправителя:",
        "Unnamed: 15": "Продавец:",
        "Unnamed: 16": "Получатель:",
        "Unnamed: 17": "Адрес получателя:",
        "Unnamed: 18": "Покупатель:",
    }

    try:
        # Читаем Excel файл через openpyxl для точного чтения чисел
        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = workbook["PL"]
        
        # Также читаем через pandas для удобства работы со структурой
        df = pd.read_excel(io.BytesIO(file_content), sheet_name="PL")
        
        # Начинаем сканирование с 1 строки (индекс 0 в pandas, но строка 2 в Excel)
        start_row = 1
        data_df = df.iloc[start_row:].copy()
        # Переменные для подсчета общих значений
        calculated_total_quantity = 0
        calculated_total_weight = 0
        calculated_total_amount = 0
        
        # Сначала собираем все данные для анализа инвойсов по контейнерам
        container_invoices = {}  # {container_number: set(invoice_numbers)}
        all_items = []  # Список всех товаров для последующей обработки
        
        # Первый проход - собираем информацию о контейнерах и инвойсах
        for idx, row in data_df.iterrows():
            container_number = row.get('Unnamed: 10')
            if pd.isna(container_number) or str(container_number).strip() == '':
                continue
            
            container_number = str(container_number).strip()
            if CON_NUMBER == container_number or CON_NUMBER is None:
                invoice_number = str(row.get('Unnamed: 11', '')).strip() if pd.notna(row.get('Unnamed: 11')) else ''
                
                if container_number not in container_invoices:
                    container_invoices[container_number] = set()
                if invoice_number:
                    container_invoices[container_number].add(invoice_number)
                
                all_items.append((idx, row))
        
        # Второй проход - обрабатываем товары с правильными ключами
        for idx, row in all_items:
            # Определяем номер строки в Excel (idx - это индекс в pandas, +1 для заголовка, +1 для 1-based индексации Excel)
            excel_row = idx + 2  # +1 для заголовка, +1 для 1-based индексации
            
            container_number = str(row.get('Unnamed: 10')).strip()
            invoice_number = str(row.get('Unnamed: 11', '')).strip() if pd.notna(row.get('Unnamed: 11')) else ''
            
            # Определяем ключ контейнера
            # Если в контейнере несколько разных инвойсов, добавляем номер инвойса к ключу
            if len(container_invoices[container_number]) > 1 and invoice_number:
                container_key = f"{container_number}_{invoice_number}"
            else:
                container_key = container_number
            
            # Читаем числовые значения напрямую из Excel через openpyxl для точности
            # В pandas "Unnamed: 4" = колонка 4 (0-based) = колонка 5 в Excel (1-based)
            # excel_row уже 1-based (строка в Excel), функция ожидает 0-based индекс строки
            quantity_places = get_precise_float_from_excel(workbook, "PL", excel_row - 1, 4)  # Unnamed: 4
            weight_brutto = get_precise_float_from_excel(workbook, "PL", excel_row - 1, 7)    # Unnamed: 7
            weight_netto = get_precise_float_from_excel(workbook, "PL", excel_row - 1, 6)     # Unnamed: 6
            amount = get_precise_float_from_excel(workbook, "PL", excel_row - 1, 9)            # Unnamed: 9
            
            # Создаем запись товара
            item = {
                "Код ТН ВЭД": str(row.get('Unnamed: 1', '')).strip()[:6] if pd.notna(row.get('Unnamed: 1')) else '',
                "Коммерческое описание товара": str(row.get('Unnamed: 2', '')).strip() if pd.notna(row.get('Unnamed: 2')) else '',
                "Признак товара, свободного от применения запретов и ограничений (всегда 1)": 1,
                "Информация об упаковке (0-БЕЗ, 1 С)":
                    0 if row.get('Unnamed: 5', '').strip() in {"NE", "NF", "NG", "PP"} else (1 if weight_brutto >= weight_netto else 0),
                #"Кол-во штук": float(row.get('Unnamed: 3', 0)) if pd.notna(row.get('Unnamed: 3')) else 0,
                "Количество грузовых мест": quantity_places,
                "Вид информации об упаковке (всегда 0)": 0,
                "Вид упаковки ": str(row.get('Unnamed: 5', '')).strip() if pd.notna(row.get('Unnamed: 5')) else '',
                "Количество упаковок": quantity_places,
                #"Нетто": float(row.get('Unnamed: 6', 0)) if pd.notna(row.get('Unnamed: 6')) else 0,
                "Номер контейнера": container_number, 
                "Вес брутто": weight_brutto,
                "Валюта": str(row.get('Unnamed: 8', '')).strip() if pd.notna(row.get('Unnamed: 8')) else '',
                "Сумма": amount,
                #"Номер инвойса": str(row.get('Unnamed: 11', '')).strip() if pd.notna(row.get('Unnamed: 11')) else '',
                #"Case No": str(row.get('Unnamed: 10', '')).strip() if pd.notna(row.get('Unnamed: 10')) else '',
            }
            
            # Добавляем товар в контейнер (используем уникальный ключ)
            if container_key not in storage.containers:
                storage.containers[container_key] = []
            
            storage.containers[container_key].append(item)
            
            sender_name = str(row.get('Unnamed: 13', '')).strip() + (f" П/П {str(row.get('Unnamed: 15', '')).strip()}" if pd.notna(row.get('Unnamed: 15')) and str(row.get('Unnamed: 15')).strip() else "")
            sender_address = str(row.get('Unnamed: 14', '')).strip()
            recipient_name = str(row.get('Unnamed: 16', '')).strip() + (f" П/П {str(row.get('Unnamed: 18', '')).strip()}" if pd.notna(row.get('Unnamed: 18')) and str(row.get('Unnamed: 18')).strip() else "")
            recipient_address = str(row.get('Unnamed: 17', '')).strip()
           
            # Сохраняем информацию об отправителе и получателе для каждого контейнера
            if container_key not in storage.container_info:
                storage.container_info[container_key] = {
                    'sender_name': sender_name,
                    'sender_address': sender_address,
                    'recipient_name': recipient_name,
                    'recipient_address': recipient_address,
                    'invoice': str(row.get('Unnamed: 11', '')).strip() if pd.notna(row.get('Unnamed: 11')) else '',
                    'date_invoice': str(row.get('Unnamed: 12', '')).strip() if pd.notna(row.get('Unnamed: 12')) else ''
                }
            
            # Также сохраняем общую информацию для совместимости
            storage.sender_name = sender_name
            storage.sender_address = sender_address
            storage.recipient_name = recipient_name
            storage.recipient_address = recipient_address

            storage.invoice = str(row.get('Unnamed: 11', '')).strip() if pd.notna(row.get('Unnamed: 11')) else ''
            storage.date_invoice = str(row.get('Unnamed: 12', '')).strip() if pd.notna(row.get('Unnamed: 12')) else ''

            # Подсчитываем общие значения
            calculated_total_quantity += item["Количество грузовых мест"]
            calculated_total_weight += item["Вес брутто"]
            calculated_total_amount += item["Сумма"]
            
            
    except Exception as e:
        return {"error": str(e)}
    # Обновляем рассчитанные значения
    storage.calc.calc_quantity = calculated_total_quantity
    storage.calc.calc_weight = calculated_total_weight
    storage.calc.calc_amount = calculated_total_amount
    # Устанавливаем значения по умолчанию для totals (можно будет обновить при необходимости)
    storage.totals.total_quantity = calculated_total_quantity
    storage.totals.total_weight = calculated_total_weight
    storage.totals.total_amount = calculated_total_amount

    # Возвращаем результат обработки
    return {"success": True, "storage": storage}

